HELLO WORLD TEST
line 2
"""
Break Time Discrepancy Analysis Engine
Compares ADP and Amazon break records and returns matched/discrepancy data.
"""

import pandas as pd
import numpy as np
import re
from datetime import datetime
from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


# -----------------------------------------------------------------
# ADP PARSER
# -----------------------------------------------------------------

def parse_adp(file) -> pd.DataFrame:
    raw = pd.read_excel(file, sheet_name=0, header=None)

    # Dynamically find the header row (contains "Time In")
    header_row = next(
        (i for i, row in raw.iterrows() if any("Time In" in str(v) for v in row)),
        2,
    )
    data = raw.iloc[header_row + 1:].copy()
    data[0] = data[0].replace("", np.nan).ffill()

    # Drop totals and blanks
    data = data[data[0].astype(str).str.strip().str.len() > 2]
    data = data[~data[0].astype(str).str.strip().str.lower().str.startswith("total")]
    data = data[data[5].notna() & data[6].notna()]

    def to_time(val):
        if pd.isna(val):
            return None
        if isinstance(val, datetime):
            return val
        try:
            return pd.to_datetime(str(val))
        except Exception:
            return None

    data["time_in"] = data[5].apply(to_time)
    data["time_out"] = data[6].apply(to_time)
    data["emp_name"] = data[0].astype(str).str.strip()

    records = []
    for emp, grp in data.groupby("emp_name", sort=False):
        grp = grp.sort_values("time_in").reset_index(drop=True)
        break_min, b_start, b_end = None, None, None
        if len(grp) >= 2:
            first_out = grp.iloc[0]["time_out"]
            second_in = grp.iloc[1]["time_in"]
            if (
                first_out is not None
                and second_in is not None
                and pd.notna(first_out)
                and pd.notna(second_in)
            ):
                delta = (second_in - first_out).total_seconds() / 60
                if 0 < delta < 120:
                    break_min = round(delta, 1)
                    b_start, b_end = first_out, second_in
        records.append(
            {
                "adp_name": emp,
                "adp_break_start": b_start,
                "adp_break_end": b_end,
                "adp_break_minutes": break_min,
            }
        )
    return pd.DataFrame(records)


# -----------------------------------------------------------------
# AMAZON PARSER
# -----------------------------------------------------------------

def parse_amazon(file) -> pd.DataFrame:
    raw = pd.read_excel(file, sheet_name=0, header=None)
    header_row = next(
        i for i, row in raw.iterrows() if any("DA Name" in str(v) for v in row)
    )
    data = raw.iloc[header_row + 1:].copy()
    data.columns = raw.iloc[header_row].tolist()
    data = data.dropna(how="all")

    col_map = {}
    for col in data.columns:
        cs = str(col).strip()
        if "DA Name" in cs:
            col_map[col] = "amz_name"
        elif "Transporter" in cs:
            col_map[col] = "transporter_id"
        elif "Break Start" in cs:
            col_map[col] = "amz_break_start"
        elif "Break End" in cs:
            col_map[col] = "amz_break_end"
        elif "Duration" in cs or "Minutes" in cs:
            col_map[col] = "amz_break_minutes"
    data = data.rename(columns=col_map)

    def to_time(val):
        if pd.isna(val):
            return None
        if isinstance(val, datetime):
            return val
        try:
            return pd.to_datetime(str(val))
        except Exception:
            return None

    data["amz_break_minutes"] = pd.to_numeric(
        data.get("amz_break_minutes", pd.Series(dtype=float)), errors="coerce"
    )
    data["amz_break_start"] = data.get(
        "amz_break_start", pd.Series(dtype=object)
    ).apply(to_time)
    data["amz_break_end"] = data.get("amz_break_end", pd.Series(dtype=object)).apply(
        to_time
    )
    data["amz_name"] = (
        data.get("amz_name", pd.Series(dtype=str)).astype(str).str.strip()
    )

    # Filter invalid rows
    data = data[data["amz_name"].str.len() > 1]
    data = data[data["amz_name"].str.lower() != "nan"]

    # Duplicates: keep highest break duration
    data = data.sort_values("amz_break_minutes", ascending=False)
    data = data.drop_duplicates(subset=["amz_name"], keep="first")

    return data[
        ["amz_name", "transporter_id", "amz_break_start", "amz_break_end", "amz_break_minutes"]
    ].reset_index(drop=True)


# -----------------------------------------------------------------
# FUZZY NAME MATCHING
# -----------------------------------------------------------------

def _split_camelcase(name: str) -> str:
    return re.sub(r"([a-z])([A-Z])", r"\1 \2", name).strip()


def _name_tokens(name: str) -> set:
    name = _split_camelcase(str(name))
    name = re.sub(r"[^a-zA-Z\s]", " ", name).lower()
    return {w for w in name.split() if len(w) > 1}


def _token_matches(ta: set, tb: set) -> int:
    return sum(1 for a in ta if any(a in b or b in a for b in tb))


def _overlap_score(name_a: str, name_b: str) -> float:
    ta, tb = _name_tokens(name_a), _name_tokens(name_b)
    if not ta or not tb:
        return 0.0
    return _token_matches(ta, tb) / max(len(ta), len(tb))


def match_employees(adp_df: pd.DataFrame, amazon_df: pd.DataFrame) -> pd.DataFrame:
    THRESHOLD = 0.33
    used_amz = set()
    matches = []

    for _, adp_row in adp_df.iterrows():
        best_score, best_amz_idx = 0, None
        for amz_idx, amz_row in amazon_df.iterrows():
            score = _overlap_score(adp_row["adp_name"], amz_row["amz_name"])
            if score > best_score:
                best_score, best_amz_idx = score, amz_idx

        if best_score >= THRESHOLD and best_amz_idx is not None:
            amz_row = amazon_df.loc[best_amz_idx]
            matches.append(
                {**adp_row.to_dict(), **amz_row.to_dict(), "match_score": round(best_score, 2)}
            )
            used_amz.add(best_amz_idx)
        else:
            matches.append(
                {
                    **adp_row.to_dict(),
                    "amz_name": None,
                    "transporter_id": None,
                    "amz_break_start": None,
                    "amz_break_end": None,
                    "amz_break_minutes": None,
                    "match_score": 0,
                }
            )

    for amz_idx, amz_row in amazon_df.iterrows():
        if amz_idx not in used_amz:
            matches.append(
                {
                    "adp_name": None,
                    "adp_break_start": None,
                    "adp_break_end": None,
                    "adp_break_minutes": None,
                    **amz_row.to_dict(),
                    "match_score": 0,
                }
            )

    return pd.DataFrame(matches)


# -----------------------------------------------------------------
# DISCREPANCY CALCULATION
# -----------------------------------------------------------------

def calculate_discrepancies(df: pd.DataFrame) -> pd.DataFrame:
    def _diff(row):
        a, b = row.get("amz_break_minutes"), row.get("adp_break_minutes")
        return round(a - b, 1) if pd.notna(a) and pd.notna(b) else None

    def _severity(row):
        a, b = row.get("amz_break_minutes"), row.get("adp_break_minutes")
        if pd.isna(a) or pd.isna(b):
            return "Missing Entry"
        d = abs(row["diff_minutes"])
        if d <= 1:    return "Match"
        elif d <= 5:  return "Minor"
        elif d <= 15: return "Moderate"
        else:          return "Major"

    def _direction(row):
        d, sev = row.get("diff_minutes"), row.get("severity")
        if sev == "Missing Entry": return "Warning Missing"
        if pd.isna(d):             return "-"
        if abs(d) <= 1:            return "Match"
        return "Amazon > ADP" if d > 0 else "ADP > Amazon"

    df["diff_minutes"] = df.apply(_diff, axis=1)
    df["severity"]     = df.apply(_severity, axis=1)
    df["direction"]    = df.apply(_direction, axis=1)
    df["needs_action"] = df["severity"].isin(["Minor", "Moderate", "Major", "Missing Entry"])
    return df


# -----------------------------------------------------------------
# CONVERSATION SCRIPTS
# -----------------------------------------------------------------

def build_script(row) -> str:
    emp  = row.get("adp_name") or row.get("amz_name") or "Employee"
    sev  = row.get("severity", "")
    a_m  = row.get("amz_break_minutes")
    b_m  = row.get("adp_break_minutes")
    diff = row.get("diff_minutes")

    if sev == "Match":
        return "No action needed - break times match."

    if sev == "Missing Entry":
        if pd.isna(a_m):
            return (
                f"Hi {emp}, I see you have a break logged in ADP ({b_m:.0f} min) "
                "but we have no record in the Amazon Delivery App. "
                "Did you log your break in the app today? Please update it before end of day."
            )
        else:
            return (
                f"Hi {emp}, we have your break in the Amazon app ({a_m:.0f} min) "
                "but there is no matching entry in ADP. "
                "Can you log your break in ADP before 5 PM today?"
            )

    longer  = "Amazon app" if diff > 0 else "ADP"
    shorter = "ADP"        if diff > 0 else "Amazon app"
    return (
        f"Hi {emp}, we noticed a discrepancy in your break times today - "
        f"your {longer} shows {abs(diff):.0f} more minutes than your {shorter}. "
        f"Amazon recorded {a_m:.0f} min and ADP recorded {b_m:.0f} min. "
        "Please review both entries and correct whichever is inaccurate by end of day."
    )


# -----------------------------------------------------------------
# FULL PIPELINE
# -----------------------------------------------------------------

def run_analysis(adp_file, amazon_file) -> pd.DataFrame:
    adp_df    = parse_adp(adp_file)
    amazon_df = parse_amazon(amazon_file)
    merged    = match_employees(adp_df, amazon_df)
    result    = calculate_discrepancies(merged)

    sev_order = {"Major": 0, "Moderate": 1, "Minor": 2, "Missing Entry": 3, "Match": 4}
    result["_sort"] = result["severity"].map(sev_order).fillna(5)
    result = result.sort_values(["_sort", "adp_name"]).reset_index(drop=True)
    result["conversation_script"] = result.apply(build_script, axis=1)
    return result


# -----------------------------------------------------------------
# EXCEL EXPORT
# -----------------------------------------------------------------

DARK_NAVY   = "1B2A4A"
MID_BLUE    = "2E5090"
LIGHT_BLUE  = "D6E4F0"
ORANGE      = "E87722"
RED_FILL    = "FDDCDC"
RED_FONT    = "C0392B"
AMBER_FILL  = "FFF3CD"
AMBER_FONT  = "856404"
YELLOW_FILL = "FFF8DC"
YELLOW_FONT = "8B6914"
GREEN_FILL  = "D5F5E3"
GREEN_FONT  = "1A7A42"
PURPLE_FILL = "EDE7F6"
PURPLE_FONT = "6A1B9A"
WHITE       = "FFFFFF"

_thin = Side(style="thin", color="BDBDBD")


def _bdr():
    return Border(left=_thin, right=_thin, top=_thin, bottom=_thin)


def _fmt_time(val):
    if val is None or (isinstance(val, float) and np.isnan(val)):
        return "-"
    try:
        return val.strftime("%-I:%M %p") if hasattr(val, "strftime") else str(val)
    except Exception:
        return str(val)


def _sev_style(sev):
    return {
        "Major":         (PatternFill("solid", fgColor=RED_FILL),    RED_FONT),
        "Moderate":      (PatternFill("solid", fgColor=AMBER_FILL),  AMBER_FONT),
        "Minor":         (PatternFill("solid", fgColor=YELLOW_FILL), YELLOW_FONT),
        "Missing Entry": (PatternFill("solid", fgColor=PURPLE_FILL), PURPLE_FONT),
    }.get(sev, (PatternFill("solid", fgColor=GREEN_FILL), GREEN_FONT))


def export_excel(df: pd.DataFrame, report_date: str = "", station: str = "DFH1") -> bytes:
    wb = Workbook()

    ws = wb.active
    ws.title = "Discrepancy Report"

    ws.merge_cells("A1:K1")
    ws["A1"] = f"  Break Time Discrepancy Report  |  Station {station}  |  {report_date}  |  Review by 5:00 PM EST"
    ws["A1"].font = Font(name="Arial", bold=True, size=13, color=WHITE)
    ws["A1"].fill = PatternFill("solid", fgColor=DARK_NAVY)
    ws["A1"].alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 32

    ws.merge_cells("A2:K2")
    ws["A2"] = "  Major (>15 min)  Moderate (6-15 min)  Minor (2-5 min)  Missing Entry  Match (<=1 min)"
    ws["A2"].font = Font(name="Arial", size=9, italic=True, color="444444")
    ws["A2"].fill = PatternFill("solid", fgColor=LIGHT_BLUE)
    ws["A2"].alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[2].height = 18

    col_headers = [
        "Employee (ADP)", "Amazon DA Name", "Amazon Break Start", "Amazon Break End",
        "Amazon Break (min)", "ADP Break Start", "ADP Break End", "ADP Break (min)",
        "Difference (min)", "Severity", "Action Required",
    ]
    for c, h in enumerate(col_headers, 1):
        cell = ws.cell(row=3, column=c, value=h)
        cell.font = Font(name="Arial", bold=True, size=10, color=WHITE)
        cell.fill = PatternFill("solid", fgColor=MID_BLUE)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = _bdr()
    ws.row_dimensions[3].height = 36

    disc_count = 0
    for r_idx, row in df.iterrows():
        er = r_idx + 4
        sev = row.get("severity", "-")
        fill, fcolor = _sev_style(sev)
        a_m  = row.get("amz_break_minutes")
        b_m  = row.get("adp_break_minutes")
        diff = row.get("diff_minutes")

        vals = [
            str(row.get("adp_name") or "-"),
            str(row.get("amz_name") or "-"),
            _fmt_time(row.get("amz_break_start")),
            _fmt_time(row.get("amz_break_end")),
            f"{a_m:.0f}" if pd.notna(a_m) else "-",
            _fmt_time(row.get("adp_break_start")),
            _fmt_time(row.get("adp_break_end")),
            f"{b_m:.0f}" if pd.notna(b_m) else "-",
            f"{diff:+.1f}" if pd.notna(diff) else "-",
            sev,
            "No action" if sev == "Match" else "Correct by 5 PM",
        ]
        if row.get("needs_action"):
            disc_count += 1
        for c, v in enumerate(vals, 1):
            cell = ws.cell(row=er, column=c, value=v)
            cell.fill = fill
            cell.border = _bdr()
            cell.alignment = Alignment(horizontal="left" if c == 1 else "center", vertical="center")
            cell.font = Font(name="Arial", size=10, color=fcolor, bold=(c in [9, 10] and sev != "Match"))
        ws.row_dimensions[er].height = 20

    footer_r = len(df) + 4
    ws.merge_cells(f"A{footer_r}:K{footer_r}")
    ws[f"A{footer_r}"] = (
        f"  SUMMARY: {disc_count} employees need correction out of {len(df)} reviewed  "
        "|  All corrections due by 5:00 PM EST"
    )
    ws[f"A{footer_r}"].font = Font(name="Arial", bold=True, size=11, color=WHITE)
    ws[f"A{footer_r}"].fill = PatternFill("solid", fgColor=ORANGE)
    ws[f"A{footer_r}"].alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[footer_r].height = 24

    for i, w in enumerate([28, 26, 14, 14, 14, 14, 14, 14, 14, 16, 18], 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A4"

    ws2 = wb.create_sheet("5PM Scripts")
    ws2.merge_cells("A1:G1")
    ws2["A1"] = f"  5 PM Correction Conversations  |  Station {station}  |  {report_date}"
    ws2["A1"].font = Font(name="Arial", bold=True, size=13, color=WHITE)
    ws2["A1"].fill = PatternFill("solid", fgColor=DARK_NAVY)
    ws2["A1"].alignment = Alignment(horizontal="left", vertical="center")
    ws2.row_dimensions[1].height = 30

    for c, h in enumerate(["#", "Employee", "Severity", "Amazon (min)", "ADP (min)", "Difference", "Script"], 1):
        cell = ws2.cell(row=2, column=c, value=h)
        cell.font = Font(name="Arial", bold=True, size=10, color=WHITE)
        cell.fill = PatternFill("solid", fgColor=MID_BLUE)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = _bdr()
    ws2.row_dimensions[2].height = 28

    issues = df[df["needs_action"]].reset_index(drop=True)
    for r_idx, row in issues.iterrows():
        er = r_idx + 3
        sev = row.get("severity", "-")
        fill, fcolor = _sev_style(sev)
        a_m  = row.get("amz_break_minutes")
        b_m  = row.get("adp_break_minutes")
        diff = row.get("diff_minutes")
        vals = [
            r_idx + 1,
            str(row.get("adp_name") or row.get("amz_name") or "-"),
            sev,
            f"{a_m:.0f} min" if pd.notna(a_m) else "-",
            f"{b_m:.0f} min" if pd.notna(b_m) else "-",
            f"{diff:+.1f} min" if pd.notna(diff) else "-",
            row.get("conversation_script", ""),
        ]
        for c, v in enumerate(vals, 1):
            cell = ws2.cell(row=er, column=c, value=v)
            cell.fill = fill
            cell.border = _bdr()
            cell.font = Font(name="Arial", size=10, color=fcolor, bold=(c in [3, 6]))
            cell.alignment = Alignment(
                horizontal="left" if c in [2, 7] else "center",
                vertical="center",
                wrap_text=True,
            )
        ws2.row_dimensions[er].height = 90

    for i, w in enumerate([5, 28, 16, 14, 14, 12, 75], 1):
        ws2.column_dimensions[get_column_letter(i)].width = w
    ws2.freeze_panes = "A3"

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()
